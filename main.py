import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import requests
from bs4 import BeautifulSoup
from PIL import ImageTk, Image
import io

# Increase the font size for visually impaired users
LARGE_FONT = ("Arial", 14)


def load_books():
    books = []

    try:
        wb = openpyxl.load_workbook('book_info.xlsx')
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            books.append(row)

        wb.close()
    except FileNotFoundError:
        messagebox.showinfo("File Not Found", "The Excel file does not exist.")

    return books


def create_excel_file(book_info):
    try:
        wb = openpyxl.load_workbook('book_info.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Author'
        ws['B1'] = 'Title'
        ws['C1'] = 'UPC'
        wb.save('book_info.xlsx')
        wb = openpyxl.load_workbook('book_info.xlsx')

    ws = wb.active
    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=book_info['Author'])
    ws.cell(row=row, column=2, value=book_info['Title'])
    ws.cell(row=row, column=3, value=book_info['UPC'])

    wb.save('book_info.xlsx')
    messagebox.showinfo("Success", "Book information added to the Excel file.")


def delete_book():
    global books_listbox  # Declare the variable as global
    selected_book = books_listbox.get(books_listbox.curselection())
    if selected_book:
        try:
            wb = openpyxl.load_workbook('book_info.xlsx')
            ws = wb.active

            row_number = 2  # Start with the first row after the header row

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == selected_book[0] and row[1] == selected_book[1] and row[2] == selected_book[2]:
                    ws.delete_rows(row_number, amount=1)
                    wb.save('book_info.xlsx')
                    messagebox.showinfo("Success", "Book information deleted from the Excel file.")
                    load_saved_books()
                    # Set selection to the previous item in the listbox
                    index = books_listbox.curselection()
                    if index:
                        books_listbox.selection_clear(index)
                        if index[0] > 0:
                            books_listbox.selection_set(index[0] - 1)
                            books_listbox.activate(index[0] - 1)
                            books_listbox.see(index[0] - 1)
                        else:
                            books_listbox.activate(0)
                            books_listbox.see(0)
                    break
                row_number += 1  # Increment the row number

            wb.close()
        except FileNotFoundError:
            messagebox.showinfo("File Not Found", "The Excel file does not exist.")


def add_new_book():
    book_info = collect_book_info()
    create_excel_file(book_info)
    load_saved_books()


def select_book(event):
    selection = books_listbox.curselection()
    if selection:
        selected_book = books_listbox.get(selection)
        messagebox.showinfo("Selected Book", f"Author: {selected_book[0]}\nTitle: {selected_book[1]}\nUPC: {selected_book[2]}")
        display_book_artwork(selected_book[2])


def load_saved_books():
    books = load_books()
    books_listbox.delete(0, tk.END)
    for book in books:
        books_listbox.insert(tk.END, book)


def scrape_book_info(upc):
    url = f"https://www.barcodelookup.com/{upc}"  # Replace with the appropriate Amazon URL for book search
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        # Find the relevant HTML elements containing the book information
        # Extract the author, title, and image URL from the HTML elements
        author_element = soup.find('span', class_='author')
        title_element = soup.find('span', class_='title')
        image_element = soup.find('img', class_='book-image')

        if author_element and title_element and image_element:
            author = author_element.text.strip()
            title = title_element.text.strip()
            image_url = image_element['src']
            return author, title, image_url

    return None, None, None


def download_book_artwork(image_url):
    if image_url:
        response = requests.get(image_url, stream=True)
        if response.status_code == 200:
            image_data = response.content
            image = Image.open(io.BytesIO(image_data))
            image = image.resize((200, 200), Image.ANTIALIAS)
            photo = ImageTk.PhotoImage(image)
            artwork_label.configure(image=photo)
            artwork_label.image = photo
            return
    artwork_label.configure(image="")
    artwork_label.image = None


def search_book():
    search_query = search_entry.get()

    if search_query:
        books = load_books()
        matching_books = []

        for book in books:
            if search_query.lower() in [str(value).lower() for value in book]:
                matching_books.append(book)

        if matching_books:
            books_listbox.delete(0, tk.END)
            for book in matching_books:
                books_listbox.insert(tk.END, book)
        else:
            messagebox.showinfo("No Results", "No books found matching the search query.")
    else:
        load_saved_books()


def collect_book_info():
    book_info = {}

    book_info['Author'] = author_entry.get()
    book_info['Title'] = title_entry.get()
    book_info['UPC'] = upc_entry.get()

    return book_info


def display_book_artwork(upc):
    author, title, image_url = scrape_book_info(upc)
    download_book_artwork(image_url)


window = tk.Tk()
window.title("Mom's Did I already read this program?")
window.geometry("800x500")  # Increase the window size

# Increase the font size for all labels
author_label = tk.Label(window, text="Author:", font=LARGE_FONT)
author_label.grid(row=0, column=0, padx=10, pady=5)
title_label = tk.Label(window, text="Title:", font=LARGE_FONT)
title_label.grid(row=1, column=0, padx=10, pady=5)
upc_label = tk.Label(window, text="UPC:", font=LARGE_FONT)
upc_label.grid(row=2, column=0, padx=10, pady=5)
search_label = tk.Label(window, text="Search:", font=LARGE_FONT)
search_label.grid(row=3, column=2, padx=10, pady=5)

# Increase the font size for all buttons
add_button = tk.Button(window, text="Add Book", command=add_new_book, font=LARGE_FONT)
add_button.grid(row=3, column=0, padx=10, pady=5)
delete_button = tk.Button(window, text="Delete Book", command=delete_book, font=LARGE_FONT)
delete_button.grid(row=3, column=1, padx=10, pady=5)
search_button = tk.Button(window, text="Search", command=search_book, font=LARGE_FONT)
search_button.grid(row=3, column=4, padx=10, pady=5)

# Increase the font size for the listbox
books_listbox = tk.Listbox(window, width=50, font=LARGE_FONT)
books_listbox.grid(row=4, column=0, columnspan=4, padx=10, pady=5)

# Increase the font size for the artwork label
artwork_label = tk.Label(window, font=LARGE_FONT)
artwork_label.grid(row=0, column=5, rowspan=5, padx=10, pady=5)

# Increase the font size for the entry fields
author_entry = tk.Entry(window, font=LARGE_FONT)
author_entry.grid(row=0, column=1, padx=10, pady=5)
title_entry = tk.Entry(window, font=LARGE_FONT)
title_entry.grid(row=1, column=1, padx=10, pady=5)
upc_entry = tk.Entry(window, font=LARGE_FONT)
upc_entry.grid(row=2, column=1, padx=10, pady=5)
search_entry = tk.Entry(window, font=LARGE_FONT)
search_entry.grid(row=3, column=3, padx=10, pady=5)

# Configure the scrollbar with a larger width and font size
scrollbar = ttk.Scrollbar(window, orient=tk.VERTICAL, command=books_listbox.yview)
scrollbar.grid(row=4, column=4, sticky="ns")
books_listbox.config(yscrollcommand=scrollbar.set, font=LARGE_FONT)

load_saved_books()

window.mainloop()

